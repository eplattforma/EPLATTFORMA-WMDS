"""
Route Reconciliation Report Export

Generates Excel reconciliation reports from the WMDS template.
Uses RouteStopInvoice as the source of truth for all route data.
"""

import os
import logging
from io import BytesIO
from datetime import date
from decimal import Decimal
from typing import Any, Optional

from openpyxl import load_workbook
from sqlalchemy import and_

from app import db
from models import (
    Shipment, RouteStop, RouteStopInvoice, Invoice, 
    CODReceipt, PODRecord, DeliveryDiscrepancy, 
    RouteReturnHandover, InvoicePostDeliveryCase, User
)
from timezone_utils import utc_now_for_db

logger = logging.getLogger(__name__)

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'templates', 'WMDS_Route_Reconciliation_Template.xlsx')


def build_route_reconciliation_dataset(route_id: int) -> Optional[dict]:
    """
    Build complete reconciliation dataset for a route.
    Uses RouteStopInvoice as the canonical source of truth.
    
    Returns dict with:
        - route: Shipment record
        - stops: List of RouteStop records
        - invoices: Dict of invoice_no -> Invoice
        - invoice_details: List of invoice detail rows
        - pod_map: Dict of invoice_no -> PODRecord
        - cod_map: Dict of invoice_no -> CODReceipt
        - discrepancy_map: Dict of invoice_no -> [DeliveryDiscrepancy]
        - return_handover_map: Dict of invoice_no -> RouteReturnHandover
        - post_delivery_cases: Dict of invoice_no -> InvoicePostDeliveryCase
        - summary: Aggregated summary data
    """
    route = db.session.get(Shipment, route_id)
    if not route:
        logger.error(f"Route {route_id} not found")
        return None
    
    stops = RouteStop.query.filter_by(shipment_id=route_id).order_by(RouteStop.seq_no).all()
    stop_map = {s.route_stop_id: s for s in stops}
    
    rsi_list = RouteStopInvoice.query.filter(
        RouteStopInvoice.route_stop_id.in_([s.route_stop_id for s in stops]),
        RouteStopInvoice.is_active == True
    ).all()
    
    invoice_nos = [rsi.invoice_no for rsi in rsi_list]
    if not invoice_nos:
        logger.warning(f"Route {route_id} has no active invoices")
        return {
            'route': route,
            'stops': stops,
            'invoices': {},
            'invoice_details': [],
            'pod_map': {},
            'cod_map': {},
            'discrepancy_map': {},
            'return_handover_map': {},
            'post_delivery_cases': {},
            'summary': _build_empty_summary(route)
        }
    
    invoices = Invoice.query.filter(Invoice.invoice_no.in_(invoice_nos)).all()
    invoice_map = {inv.invoice_no: inv for inv in invoices}
    
    pod_map = _build_pod_map(route_id, invoice_nos)
    cod_map = _build_cod_map(route_id, invoice_nos)
    discrepancy_map = _build_discrepancy_map(invoice_nos)
    return_handover_map = _build_return_handover_map(route_id, invoice_nos)
    post_delivery_cases = _build_post_delivery_cases_map(invoice_nos)
    
    invoice_details = []
    for rsi in rsi_list:
        stop = stop_map.get(rsi.route_stop_id)
        invoice = invoice_map.get(rsi.invoice_no)
        if not stop or not invoice:
            continue
            
        pod = pod_map.get(rsi.invoice_no)
        cod = cod_map.get(rsi.invoice_no)
        discrepancies = discrepancy_map.get(rsi.invoice_no, [])
        return_handover = return_handover_map.get(rsi.invoice_no)
        pdc = post_delivery_cases.get(rsi.invoice_no)
        
        expected_amount = rsi.expected_amount if rsi.expected_amount else invoice.total_grand
        received_amount = cod.received_amount if cod else None
        
        has_open_discrepancy = any(d.status not in ('CLOSED', 'RESOLVED') for d in discrepancies)
        has_open_pdc = pdc and pdc.status not in ('CLOSED', 'RETURN_TO_STOCK')
        
        detail = {
            'route_id': route_id,
            'stop_seq': float(stop.seq_no) if stop.seq_no else 0,
            'customer_code': invoice.customer_code,
            'invoice_no': rsi.invoice_no,
            'payment_type': rsi.expected_payment_method or _infer_payment_type(invoice),
            'expected_amount': float(expected_amount) if expected_amount else 0,
            'received_amount': float(received_amount) if received_amount else None,
            'variance': float(received_amount - expected_amount) if received_amount and expected_amount else None,
            'has_payment_evidence': cod is not None,
            'has_pod_evidence': pod is not None,
            'has_receipt_ack': bool(cod and cod.ps365_receipt_id),
            'cheque_no': cod.cheque_number if cod else None,
            'cheque_date': cod.cheque_date if cod else None,
            'online_ref': None,
            'pod_status': rsi.status or 'PENDING',
            'discrepancy_open': has_open_discrepancy,
            'post_delivery_case_open': has_open_pdc,
            'credit_note_required': any(d.needs_credit_note for d in discrepancies if hasattr(d, 'needs_credit_note')),
            'credit_note_no': None,
            'credit_note_amount': None,
            'ps365_synced': bool(cod and cod.ps365_synced_at),
            'online_allocated': None,
            'exception_flag': _has_exception(rsi, pod, cod, discrepancies, pdc),
            'exception_notes': _build_exception_notes(rsi, pod, cod, discrepancies, pdc)
        }
        invoice_details.append(detail)
    
    invoice_details.sort(key=lambda x: (x['stop_seq'], x['invoice_no']))
    
    summary = _build_summary(route, stops, invoice_details, cod_map)
    
    return {
        'route': route,
        'stops': stops,
        'invoices': invoice_map,
        'invoice_details': invoice_details,
        'pod_map': pod_map,
        'cod_map': cod_map,
        'discrepancy_map': discrepancy_map,
        'return_handover_map': return_handover_map,
        'post_delivery_cases': post_delivery_cases,
        'summary': summary
    }


def _build_pod_map(route_id: int, invoice_nos: list) -> dict:
    """Build invoice_no -> PODRecord map by exploding JSON arrays"""
    pod_map = {}
    pods = PODRecord.query.filter_by(route_id=route_id).all()
    for pod in pods:
        if pod.invoice_nos:
            for inv_no in pod.invoice_nos:
                if inv_no in invoice_nos:
                    pod_map[inv_no] = pod
    return pod_map


def _build_cod_map(route_id: int, invoice_nos: list) -> dict:
    """Build invoice_no -> CODReceipt map by exploding JSON arrays"""
    cod_map = {}
    cods = CODReceipt.query.filter_by(route_id=route_id).all()
    for cod in cods:
        if cod.invoice_nos:
            for inv_no in cod.invoice_nos:
                if inv_no in invoice_nos:
                    cod_map[inv_no] = cod
    return cod_map


def _build_discrepancy_map(invoice_nos: list) -> dict:
    """Build invoice_no -> [DeliveryDiscrepancy] map"""
    discrepancy_map = {inv_no: [] for inv_no in invoice_nos}
    discrepancies = DeliveryDiscrepancy.query.filter(
        DeliveryDiscrepancy.invoice_no.in_(invoice_nos)
    ).all()
    for d in discrepancies:
        discrepancy_map[d.invoice_no].append(d)
    return discrepancy_map


def _build_return_handover_map(route_id: int, invoice_nos: list) -> dict:
    """Build invoice_no -> RouteReturnHandover map"""
    handovers = RouteReturnHandover.query.filter(
        RouteReturnHandover.route_id == route_id,
        RouteReturnHandover.invoice_no.in_(invoice_nos)
    ).all()
    return {h.invoice_no: h for h in handovers}


def _build_post_delivery_cases_map(invoice_nos: list) -> dict:
    """Build invoice_no -> InvoicePostDeliveryCase map (most recent open case)"""
    cases = InvoicePostDeliveryCase.query.filter(
        InvoicePostDeliveryCase.invoice_no.in_(invoice_nos)
    ).order_by(InvoicePostDeliveryCase.created_at.desc()).all()
    case_map = {}
    for c in cases:
        if c.invoice_no not in case_map:
            case_map[c.invoice_no] = c
    return case_map


def _infer_payment_type(invoice: Invoice) -> str:
    """Infer payment type from invoice or customer data"""
    return 'CASH'


def _has_exception(rsi, pod, cod, discrepancies, pdc) -> bool:
    """Determine if invoice has any exception requiring attention"""
    if rsi.status in ('DELIVERED', 'PARTIAL') and not pod:
        return True
    payment_type = rsi.expected_payment_method or ''
    if payment_type.upper() in ('CASH', 'DAY CHEQUE', 'POST DATED CHQ') and not cod:
        if rsi.status in ('DELIVERED', 'PARTIAL'):
            return True
    if any(d.status not in ('CLOSED', 'RESOLVED') for d in discrepancies):
        return True
    if pdc and pdc.status not in ('CLOSED', 'RETURN_TO_STOCK'):
        return True
    return False


def _build_exception_notes(rsi, pod, cod, discrepancies, pdc) -> str:
    """Build exception notes string"""
    notes = []
    if rsi.status in ('DELIVERED', 'PARTIAL') and not pod:
        notes.append('Missing POD')
    if rsi.expected_payment_method and rsi.expected_payment_method.upper() in ('CASH', 'DAY CHEQUE', 'POST DATED CHQ'):
        if not cod and rsi.status in ('DELIVERED', 'PARTIAL'):
            notes.append('Missing payment evidence')
    open_discrepancies = [d for d in discrepancies if d.status not in ('CLOSED', 'RESOLVED')]
    if open_discrepancies:
        notes.append(f'{len(open_discrepancies)} open discrepancy')
    if pdc and pdc.status not in ('CLOSED', 'RETURN_TO_STOCK'):
        notes.append('Post-delivery case open')
    return '; '.join(notes)


def _build_empty_summary(route: Shipment) -> dict:
    """Build empty summary for routes with no invoices"""
    return {
        'route_id': route.id,
        'delivery_date': route.delivery_date,
        'driver': route.driver_name,
        'vehicle': route.route_name,
        'reconciliation_status': route.reconciliation_status,
        'settlement_status': None,
        'stops_planned': 0,
        'stops_completed': 0,
        'invoices_planned': 0,
        'delivered': 0,
        'failed': 0,
        'partial': 0,
        'settlement_by_type': {}
    }


def _build_summary(route: Shipment, stops: list, invoice_details: list, cod_map: dict) -> dict:
    """Build summary statistics from invoice details"""
    delivered = sum(1 for d in invoice_details if d['pod_status'] == 'DELIVERED')
    failed = sum(1 for d in invoice_details if d['pod_status'] == 'FAILED')
    partial = sum(1 for d in invoice_details if d['pod_status'] == 'PARTIAL')
    
    completed_stops = set()
    for d in invoice_details:
        if d['pod_status'] in ('DELIVERED', 'FAILED', 'PARTIAL'):
            completed_stops.add(d['stop_seq'])
    
    settlement_by_type = {}
    for payment_type in ['CASH', 'DAY CHEQUE', 'POST DATED CHQ', 'CREDIT', 'ONLINE']:
        type_invoices = [d for d in invoice_details if d['payment_type'] == payment_type]
        expected = sum(d['expected_amount'] for d in type_invoices)
        received = sum(d['received_amount'] or 0 for d in type_invoices)
        evidence_count = sum(1 for d in type_invoices if d['has_payment_evidence'])
        settlement_by_type[payment_type] = {
            'expected': expected,
            'received': received,
            'variance': received - expected,
            'evidence_count': evidence_count
        }
    
    return {
        'route_id': route.id,
        'delivery_date': route.delivery_date,
        'driver': route.driver_name,
        'vehicle': route.route_name,
        'reconciliation_status': route.reconciliation_status,
        'settlement_status': None,
        'stops_planned': len(stops),
        'stops_completed': len(completed_stops),
        'invoices_planned': len(invoice_details),
        'delivered': delivered,
        'failed': failed,
        'partial': partial,
        'settlement_by_type': settlement_by_type
    }


def render_excel(dataset: dict) -> BytesIO:
    """
    Render the reconciliation report Excel file from dataset.
    
    Populates:
    - Summary sheet: Route header info (formulas calculate counts)
    - Invoice Detail sheet: One row per invoice
    - StopSummary: Uses formulas from Invoice Detail
    - PostDated Register: Cheque details
    """
    wb = load_workbook(TEMPLATE_PATH)
    
    _fill_summary_sheet(wb['Summary'], dataset)
    _fill_invoice_detail_sheet(wb['Invoice Detail'], dataset)
    _fill_stop_summary_sheet(wb['StopSummary'], dataset)
    _fill_postdated_sheet(wb['PostDated Register'], dataset)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _fill_summary_sheet(ws, dataset: dict):
    """Fill Summary sheet header fields"""
    summary = dataset['summary']
    route = dataset['route']
    
    ws['B3'] = summary['route_id']
    ws['B4'] = summary['delivery_date'].strftime('%Y-%m-%d') if summary['delivery_date'] else ''
    ws['B5'] = summary['driver'] or ''
    ws['B6'] = summary['vehicle'] or ''
    ws['B7'] = summary['reconciliation_status'] or ''
    ws['B8'] = summary['settlement_status'] or ''
    ws['B9'] = route.notes if hasattr(route, 'notes') and route.notes else ''


def _fill_invoice_detail_sheet(ws, dataset: dict):
    """Fill Invoice Detail sheet with one row per invoice"""
    invoice_details = dataset['invoice_details']
    
    start_row = 4
    for idx, detail in enumerate(invoice_details):
        row = start_row + idx
        ws.cell(row=row, column=1, value=detail['route_id'])
        ws.cell(row=row, column=2, value=detail['stop_seq'])
        ws.cell(row=row, column=3, value=detail['customer_code'])
        ws.cell(row=row, column=4, value=detail['invoice_no'])
        ws.cell(row=row, column=5, value=detail['payment_type'])
        ws.cell(row=row, column=6, value=detail['expected_amount'])
        ws.cell(row=row, column=7, value=detail['received_amount'])
        ws.cell(row=row, column=9, value=detail['has_payment_evidence'])
        ws.cell(row=row, column=10, value=detail['has_pod_evidence'])
        ws.cell(row=row, column=11, value=detail['has_receipt_ack'])
        ws.cell(row=row, column=12, value=detail['cheque_no'])
        if detail['cheque_date']:
            ws.cell(row=row, column=13, value=detail['cheque_date'])
        ws.cell(row=row, column=14, value=detail['online_ref'])
        ws.cell(row=row, column=15, value=detail['pod_status'])
        ws.cell(row=row, column=16, value=detail['discrepancy_open'])
        ws.cell(row=row, column=17, value=detail['post_delivery_case_open'])
        ws.cell(row=row, column=18, value=detail['credit_note_required'])
        ws.cell(row=row, column=19, value=detail['credit_note_no'])
        ws.cell(row=row, column=20, value=detail['credit_note_amount'])
        ws.cell(row=row, column=21, value=detail['ps365_synced'])
        ws.cell(row=row, column=22, value=detail['online_allocated'])
        ws.cell(row=row, column=23, value=detail['exception_flag'])
        ws.cell(row=row, column=24, value=detail['exception_notes'])


def _fill_stop_summary_sheet(ws, dataset: dict):
    """Fill StopSummary sheet with stop data (formulas handle calculations)"""
    stops = dataset['stops']
    invoices = dataset['invoices']
    
    start_row = 4
    for idx, stop in enumerate(stops):
        row = start_row + idx
        ws.cell(row=row, column=1, value=float(stop.seq_no) if stop.seq_no else idx + 1)
        ws.cell(row=row, column=2, value=stop.route_stop_id)
        ws.cell(row=row, column=3, value=stop.customer_code)
        ws.cell(row=row, column=4, value=stop.stop_name)
        ws.cell(row=row, column=5, value=stop.stop_addr)


def _fill_postdated_sheet(ws, dataset: dict):
    """Fill PostDated Register sheet with post-dated cheque details"""
    invoice_details = dataset['invoice_details']
    postdated_entries = [
        d for d in invoice_details 
        if d['payment_type'] == 'POST DATED CHQ' and d['cheque_no']
    ]
    
    start_row = 4
    for idx, entry in enumerate(postdated_entries):
        row = start_row + idx
        ws.cell(row=row, column=1, value=entry['cheque_no'])
        ws.cell(row=row, column=3, value=entry['customer_code'])
        ws.cell(row=row, column=4, value=entry['invoice_no'])
        ws.cell(row=row, column=5, value=entry['received_amount'])
        if entry['cheque_date']:
            ws.cell(row=row, column=6, value=entry['cheque_date'])
        ws.cell(row=row, column=8, value=entry['route_id'])
        ws.cell(row=row, column=9, value=entry['stop_seq'])


def generate_route_reconciliation_excel(route_id: int) -> Optional[BytesIO]:
    """
    Main entry point: Generate reconciliation Excel for a route.
    
    Returns BytesIO with Excel content, or None if route not found.
    """
    dataset = build_route_reconciliation_dataset(route_id)
    if not dataset:
        return None
    return render_excel(dataset)
