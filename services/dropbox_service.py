import os
import logging
import time
import hashlib
import base64
import requests
from datetime import datetime, timedelta
from cryptography.fernet import Fernet, InvalidToken
from app import db
from models import ExternalAuthCredential, ExternalFileSyncLog
from timezone_utils import get_utc_now

logger = logging.getLogger(__name__)

PROVIDER = 'dropbox'
DROPBOX_AUTH_URL = 'https://www.dropbox.com/oauth2/authorize'
DROPBOX_TOKEN_URL_API = 'https://api.dropboxapi.com/oauth2/token'
DROPBOX_ACCOUNT_URL = 'https://api.dropboxapi.com/2/users/get_current_account'
DROPBOX_METADATA_URL = 'https://api.dropboxapi.com/2/files/get_metadata'
DROPBOX_DOWNLOAD_URL = 'https://content.dropboxapi.com/2/files/download'

TOKEN_EXPIRY_BUFFER = timedelta(minutes=5)
MAX_RETRIES = 2
RETRY_BACKOFF = [1, 3]
SYNC_LOCK_MINUTES = 15

VALID_STATUSES = (
    'success', 'success_no_change', 'auth_error', 'download_error',
    'parse_error', 'config_error', 'running', 'skipped_concurrent',
)


def _get_fernet():
    secret = os.environ.get('SESSION_SECRET', '')
    if not secret:
        raise ValueError("SESSION_SECRET is required for token encryption")
    key = base64.urlsafe_b64encode(hashlib.sha256(secret.encode()).digest())
    return Fernet(key)


def _encrypt_token(plaintext):
    if not plaintext:
        return plaintext
    f = _get_fernet()
    return 'fernet:' + f.encrypt(plaintext.encode()).decode()


def _decrypt_token(stored):
    if not stored:
        return stored
    if stored.startswith('fernet:'):
        try:
            f = _get_fernet()
            return f.decrypt(stored[7:].encode()).decode()
        except (InvalidToken, Exception) as e:
            logger.error(f"Token decryption failed: {e}")
            return None
    return stored


def _get_config():
    app_key = os.environ.get('DROPBOX_APP_KEY', '').strip()
    app_secret = os.environ.get('DROPBOX_APP_SECRET', '').strip()
    redirect_uri = os.environ.get('DROPBOX_REDIRECT_URI', '').strip()
    file_path = os.environ.get('DROPBOX_FILE_PATH', '').strip()
    return {
        'app_key': app_key,
        'app_secret': app_secret,
        'redirect_uri': redirect_uri,
        'file_path': file_path,
        'configured': bool(app_key and app_secret),
    }


def get_dropbox_credentials():
    return db.session.query(ExternalAuthCredential).filter_by(
        provider=PROVIDER
    ).order_by(ExternalAuthCredential.id.desc()).first()


def build_dropbox_authorize_url(state_token):
    config = _get_config()
    if not config['configured']:
        raise ValueError("Dropbox app key/secret not configured")
    params = {
        'client_id': config['app_key'],
        'response_type': 'code',
        'token_access_type': 'offline',
        'state': state_token,
    }
    if config['redirect_uri']:
        params['redirect_uri'] = config['redirect_uri']
    qs = '&'.join(f"{k}={requests.utils.quote(str(v))}" for k, v in params.items())
    return f"{DROPBOX_AUTH_URL}?{qs}"


def exchange_code_for_tokens(code):
    config = _get_config()
    if not config['configured']:
        raise ValueError("Dropbox app key/secret not configured")

    data = {
        'code': code,
        'grant_type': 'authorization_code',
        'client_id': config['app_key'],
        'client_secret': config['app_secret'],
    }
    if config['redirect_uri']:
        data['redirect_uri'] = config['redirect_uri']

    logger.info("Exchanging authorization code for tokens")
    resp = requests.post(DROPBOX_TOKEN_URL_API, data=data, timeout=30)
    if resp.status_code != 200:
        logger.error(f"Token exchange failed: {resp.status_code}")
        raise ValueError(f"Token exchange failed (HTTP {resp.status_code})")

    token_data = resp.json()
    now = get_utc_now()

    db.session.query(ExternalAuthCredential).filter(
        ExternalAuthCredential.provider == PROVIDER,
        ExternalAuthCredential.status == 'active'
    ).update({'status': 'replaced'})

    cred = get_dropbox_credentials()
    if cred and cred.status in ('disconnected', 'auth_error', 'replaced'):
        pass
    else:
        cred = ExternalAuthCredential(provider=PROVIDER)
        db.session.add(cred)

    cred.refresh_token = _encrypt_token(token_data['refresh_token'])
    cred.access_token = token_data['access_token']
    expires_in = token_data.get('expires_in', 14400)
    cred.access_token_expires_at = now + timedelta(seconds=expires_in)
    cred.scope_text = token_data.get('scope', '')
    cred.dropbox_account_id = token_data.get('account_id', '')
    cred.status = 'active'
    cred.last_auth_at = now
    cred.last_error = None

    account_info = _fetch_account_info(token_data['access_token'])
    if account_info:
        cred.dropbox_email = account_info.get('email', '')
        cred.account_label = account_info.get('name', {}).get('display_name', '')

    db.session.commit()
    logger.info(f"Dropbox connected: {cred.dropbox_email or cred.dropbox_account_id}")
    return cred


def _fetch_account_info(access_token):
    try:
        resp = requests.post(
            DROPBOX_ACCOUNT_URL,
            headers={'Authorization': f'Bearer {access_token}'},
            timeout=15
        )
        if resp.status_code == 200:
            return resp.json()
    except Exception as e:
        logger.warning(f"Could not fetch Dropbox account info: {e}")
    return None


def refresh_dropbox_access_token(force=False):
    config = _get_config()
    cred = get_dropbox_credentials()
    if not cred or cred.status != 'active':
        raise ValueError("No active Dropbox credential found — reconnect required")

    now = get_utc_now()
    if not force and cred.access_token and cred.access_token_expires_at:
        if cred.access_token_expires_at > now + TOKEN_EXPIRY_BUFFER:
            return cred.access_token

    raw_refresh = _decrypt_token(cred.refresh_token)
    if not raw_refresh:
        cred.status = 'auth_error'
        cred.last_error = 'Refresh token missing or corrupt'
        db.session.commit()
        raise ValueError("Refresh token unavailable — reconnect required")

    logger.info("Refreshing Dropbox access token")
    data = {
        'grant_type': 'refresh_token',
        'refresh_token': raw_refresh,
        'client_id': config['app_key'],
        'client_secret': config['app_secret'],
    }

    resp = requests.post(DROPBOX_TOKEN_URL_API, data=data, timeout=30)
    if resp.status_code != 200:
        error_msg = f"Token refresh failed (HTTP {resp.status_code})"
        logger.error(error_msg)
        cred.status = 'auth_error'
        cred.last_error = error_msg
        db.session.commit()
        raise ValueError("Token refresh failed — reconnect required")

    token_data = resp.json()
    cred.access_token = token_data['access_token']
    expires_in = token_data.get('expires_in', 14400)
    cred.access_token_expires_at = now + timedelta(seconds=expires_in)
    cred.last_refresh_at = now
    cred.last_error = None
    db.session.commit()
    logger.info("Dropbox access token refreshed successfully")
    return cred.access_token


def get_valid_dropbox_access_token():
    return refresh_dropbox_access_token(force=False)


def _dropbox_api_call(method, url, headers=None, **kwargs):
    token = get_valid_dropbox_access_token()
    hdrs = dict(headers or {})
    hdrs['Authorization'] = f'Bearer {token}'

    for attempt in range(MAX_RETRIES):
        try:
            resp = requests.request(method, url, headers=hdrs, timeout=60, **kwargs)
            if resp.status_code == 401 and attempt == 0:
                logger.warning("Dropbox 401 — forcing token refresh and retrying")
                token = refresh_dropbox_access_token(force=True)
                hdrs['Authorization'] = f'Bearer {token}'
                continue
            if resp.status_code in (429, 500, 502, 503) and attempt == 0:
                wait = RETRY_BACKOFF[0]
                retry_after = resp.headers.get('Retry-After')
                if retry_after:
                    try:
                        wait = min(int(retry_after), 10)
                    except ValueError:
                        pass
                logger.warning(f"Dropbox {resp.status_code} — retrying in {wait}s")
                time.sleep(wait)
                continue
            return resp
        except requests.exceptions.RequestException as e:
            if attempt < MAX_RETRIES - 1:
                wait = RETRY_BACKOFF[attempt] if attempt < len(RETRY_BACKOFF) else 3
                logger.warning(f"Dropbox request error (attempt {attempt+1}): {e}, retrying in {wait}s")
                time.sleep(wait)
            else:
                raise
    return resp


def get_dropbox_file_metadata(path):
    import json
    resp = _dropbox_api_call(
        'POST', DROPBOX_METADATA_URL,
        headers={'Content-Type': 'application/json'},
        json={'path': path}
    )
    if resp.status_code == 409:
        error_detail = ''
        try:
            error_detail = resp.json().get('error_summary', '')
        except Exception:
            pass
        raise ValueError(f"File not found or inaccessible: {error_detail}")
    if resp.status_code != 200:
        raise ValueError(f"Metadata fetch failed (HTTP {resp.status_code})")
    return resp.json()


def download_dropbox_file_bytes(path):
    import json
    resp = _dropbox_api_call(
        'POST', DROPBOX_DOWNLOAD_URL,
        headers={'Dropbox-API-Arg': json.dumps({'path': path})}
    )
    if resp.status_code != 200:
        raise ValueError(f"Download failed (HTTP {resp.status_code})")

    metadata = {}
    api_result = resp.headers.get('Dropbox-API-Result', '')
    if api_result:
        try:
            metadata = json.loads(api_result)
        except Exception:
            pass

    return resp.content, metadata


def get_dropbox_status():
    config = _get_config()
    cred = get_dropbox_credentials()

    last_sync = db.session.query(ExternalFileSyncLog).filter_by(
        provider=PROVIDER
    ).order_by(ExternalFileSyncLog.started_at.desc()).first()

    last_success = db.session.query(ExternalFileSyncLog).filter(
        ExternalFileSyncLog.provider == PROVIDER,
        ExternalFileSyncLog.status.in_(['success', 'success_no_change'])
    ).order_by(ExternalFileSyncLog.started_at.desc()).first()

    recent_logs = db.session.query(ExternalFileSyncLog).filter_by(
        provider=PROVIDER
    ).order_by(ExternalFileSyncLog.started_at.desc()).limit(50).all()

    return {
        'configured': config['configured'],
        'file_path': config['file_path'],
        'connected': cred is not None and cred.status == 'active',
        'status': cred.status if cred else 'not_connected',
        'email': cred.dropbox_email if cred else None,
        'account_label': cred.account_label if cred else None,
        'last_auth_at': cred.last_auth_at if cred else None,
        'last_refresh_at': cred.last_refresh_at if cred else None,
        'last_error': cred.last_error if cred else None,
        'last_sync': {
            'status': last_sync.status if last_sync else None,
            'started_at': last_sync.started_at if last_sync else None,
            'finished_at': last_sync.finished_at if last_sync else None,
            'rows_imported': last_sync.rows_imported if last_sync else 0,
            'file_revision': last_sync.file_revision if last_sync else None,
            'file_modified_at': last_sync.file_modified_at if last_sync else None,
            'error_message': last_sync.error_message if last_sync else None,
        },
        'last_success': {
            'started_at': last_success.started_at if last_success else None,
            'rows_imported': last_success.rows_imported if last_success else 0,
            'file_revision': last_success.file_revision if last_success else None,
            'file_modified_at': last_success.file_modified_at if last_success else None,
        },
        'sync_history': recent_logs,
    }


def _check_sync_lock():
    cutoff = get_utc_now() - timedelta(minutes=SYNC_LOCK_MINUTES)
    running = db.session.query(ExternalFileSyncLog).filter(
        ExternalFileSyncLog.provider == PROVIDER,
        ExternalFileSyncLog.status == 'running',
        ExternalFileSyncLog.started_at > cutoff,
    ).first()
    return running


def sync_dropbox_file(file_processor=None, skip_unchanged=True):
    config = _get_config()
    file_path = config['file_path'] or '(not configured)'

    log = ExternalFileSyncLog(
        provider=PROVIDER,
        file_path=file_path,
        status='running',
        started_at=get_utc_now(),
    )
    db.session.add(log)
    db.session.commit()
    sync_log_id = log.id
    logger.info(f"[sync:{sync_log_id}] Dropbox sync started for {file_path}")

    try:
        if not config['configured']:
            raise _SyncError("Dropbox app key/secret not configured in environment", 'config_error')
        if not config['file_path']:
            raise _SyncError("DROPBOX_FILE_PATH not configured", 'config_error')

        existing_running = db.session.query(ExternalFileSyncLog).filter(
            ExternalFileSyncLog.provider == PROVIDER,
            ExternalFileSyncLog.status == 'running',
            ExternalFileSyncLog.id != sync_log_id,
            ExternalFileSyncLog.started_at > get_utc_now() - timedelta(minutes=SYNC_LOCK_MINUTES),
        ).first()
        if existing_running:
            log.status = 'skipped_concurrent'
            log.error_message = 'Another sync is already running'
            log.finished_at = get_utc_now()
            db.session.commit()
            logger.info(f"[sync:{sync_log_id}] Skipped — concurrent sync in progress (id={existing_running.id})")
            return log

        cred = get_dropbox_credentials()
        if not cred or cred.status != 'active':
            raise _SyncError("Dropbox not connected — please connect from the admin UI", 'auth_error')

        logger.info(f"[sync:{sync_log_id}] Fetching metadata")
        metadata = get_dropbox_file_metadata(file_path)
        log.file_name = metadata.get('name', '')
        log.file_revision = metadata.get('rev', '')
        modified_str = metadata.get('server_modified', '')
        if modified_str:
            try:
                log.file_modified_at = datetime.fromisoformat(modified_str.replace('Z', '+00:00'))
            except Exception:
                pass
        content_hash = metadata.get('content_hash', '')
        log.metadata_json = {
            'size': metadata.get('size'),
            'content_hash': content_hash,
        }
        db.session.commit()

        if skip_unchanged and content_hash:
            prev_success = db.session.query(ExternalFileSyncLog).filter(
                ExternalFileSyncLog.provider == PROVIDER,
                ExternalFileSyncLog.status == 'success',
                ExternalFileSyncLog.id != sync_log_id,
            ).order_by(ExternalFileSyncLog.started_at.desc()).first()
            if prev_success and prev_success.metadata_json:
                prev_hash = prev_success.metadata_json.get('content_hash', '')
                if prev_hash and prev_hash == content_hash:
                    log.status = 'success_no_change'
                    log.rows_imported = prev_success.rows_imported
                    log.finished_at = get_utc_now()
                    db.session.commit()
                    logger.info(f"[sync:{sync_log_id}] File unchanged (hash match) — skipped import")
                    return log

        logger.info(f"[sync:{sync_log_id}] Downloading file")
        file_bytes, dl_metadata = download_dropbox_file_bytes(file_path)
        logger.info(f"[sync:{sync_log_id}] Downloaded {len(file_bytes)} bytes")

        rows_imported = 0
        if file_processor:
            rows_imported = file_processor(file_bytes, metadata)
        else:
            rows_imported = _cost_import_processor(file_bytes, sync_log_id)

        log.status = 'success'
        log.rows_imported = rows_imported
        log.finished_at = get_utc_now()
        db.session.commit()
        duration = (log.finished_at - log.started_at).total_seconds()
        logger.info(f"[sync:{sync_log_id}] Complete: {rows_imported} rows in {duration:.1f}s")
        return log

    except _SyncError as e:
        log.status = e.sync_status
        log.error_message = str(e)
        log.finished_at = get_utc_now()
        db.session.commit()
        logger.error(f"[sync:{sync_log_id}] {e.sync_status}: {e}")
        raise ValueError(str(e))

    except ValueError as e:
        error_msg = str(e)
        if 'auth' in error_msg.lower() or '401' in error_msg or 'token' in error_msg.lower() or 'reconnect' in error_msg.lower():
            log.status = 'auth_error'
        elif 'not found' in error_msg.lower() or 'not_found' in error_msg.lower() or '409' in error_msg:
            log.status = 'download_error'
        elif 'parse' in error_msg.lower() or 'worksheet' in error_msg.lower() or 'column' in error_msg.lower():
            log.status = 'parse_error'
        else:
            log.status = 'download_error'
        log.error_message = error_msg
        log.finished_at = get_utc_now()
        db.session.commit()
        logger.error(f"[sync:{sync_log_id}] {log.status}: {error_msg}")
        raise

    except Exception as e:
        log.status = 'download_error'
        log.error_message = str(e)[:500]
        log.finished_at = get_utc_now()
        db.session.commit()
        logger.error(f"[sync:{sync_log_id}] download_error: {e}")
        raise


class _SyncError(Exception):
    def __init__(self, message, sync_status='download_error'):
        super().__init__(message)
        self.sync_status = sync_status


def _cost_import_processor(file_bytes, sync_log_id=None):
    import openpyxl
    from io import BytesIO
    from decimal import Decimal, InvalidOperation
    from models import DwItem

    tag = f"[sync:{sync_log_id}]" if sync_log_id else "[sync]"

    try:
        workbook = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        raise _SyncError(f"Cannot open workbook: {e}", 'parse_error')

    worksheet = workbook.active
    if worksheet is None:
        raise _SyncError("No active worksheet found in workbook", 'parse_error')

    logger.info(f"{tag} Parsing worksheet: {worksheet.title}")

    header_row = None
    item_code_col = None
    cost_col = None

    for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10, values_only=False), start=1):
        cells = [(c.value, c.column - 1) for c in row if c.value]
        for val, col_idx in cells:
            cell_lower = str(val).strip().lower()
            if item_code_col is None and any(kw in cell_lower for kw in ['item code', 'item_code', 'itemcode', 'item no', 'item_no']):
                item_code_col = col_idx
                header_row = row_idx
            if cost_col is None and any(kw in cell_lower for kw in ['cost', 'cost price', 'cost_price', 'costprice', 'unit cost']):
                cost_col = col_idx
                header_row = row_idx
        if item_code_col is not None and cost_col is not None:
            break

    if item_code_col is None or cost_col is None:
        sample_rows = []
        for row in worksheet.iter_rows(min_row=1, max_row=3, values_only=True):
            sample_rows.append([str(c)[:40] if c else '' for c in (row or [])[:10]])
        raise _SyncError(
            f"Cannot find required columns. Need 'Item Code' and 'Cost' columns. "
            f"Found item_code_col={'yes' if item_code_col is not None else 'NO'}, "
            f"cost_col={'yes' if cost_col is not None else 'NO'}. "
            f"First rows: {sample_rows}",
            'parse_error'
        )

    logger.info(f"{tag} Header at row {header_row}: item_code=col{item_code_col}, cost=col{cost_col}")

    data_start = header_row + 1
    rows_read = 0
    rows_matched = 0
    rows_updated = 0
    rows_skipped_blank_cost = 0
    rows_skipped_no_code = 0
    parse_errors = 0
    unmatched_codes = []

    all_item_codes = set(
        code for (code,) in db.session.query(DwItem.item_code_365).all()
    )
    logger.info(f"{tag} Loaded {len(all_item_codes)} existing item codes from ps_items_dw")

    updates = {}

    for row in worksheet.iter_rows(min_row=data_start, values_only=True):
        if not row:
            continue
        rows_read += 1

        item_code_raw = row[item_code_col] if item_code_col < len(row) else None
        cost_raw = row[cost_col] if cost_col < len(row) else None

        if not item_code_raw or str(item_code_raw).strip() == '':
            rows_skipped_no_code += 1
            continue

        item_code = str(item_code_raw).strip()

        if cost_raw is None or str(cost_raw).strip() == '':
            rows_skipped_blank_cost += 1
            continue

        try:
            cost_value = Decimal(str(cost_raw).strip())
        except (InvalidOperation, ValueError, TypeError):
            parse_errors += 1
            continue

        if item_code in all_item_codes:
            rows_matched += 1
            updates[item_code] = cost_value
        else:
            unmatched_codes.append(item_code)

    workbook.close()

    if not rows_read:
        raise _SyncError(f"No data rows found after header (row {header_row})", 'parse_error')

    batch_size = 500
    item_codes_list = list(updates.keys())
    for i in range(0, len(item_codes_list), batch_size):
        batch_codes = item_codes_list[i:i + batch_size]
        for code in batch_codes:
            item = db.session.query(DwItem).get(code)
            if item:
                if item.cost_price != updates[code]:
                    item.cost_price = updates[code]
                    rows_updated += 1

    db.session.commit()

    log_entry = db.session.query(ExternalFileSyncLog).get(sync_log_id) if sync_log_id else None
    if log_entry:
        md = log_entry.metadata_json or {}
        md['rows_read'] = rows_read
        md['rows_matched'] = rows_matched
        md['rows_updated'] = rows_updated
        md['rows_skipped_blank_cost'] = rows_skipped_blank_cost
        md['rows_skipped_no_code'] = rows_skipped_no_code
        md['parse_errors'] = parse_errors
        md['unmatched_count'] = len(unmatched_codes)
        md['unmatched_codes'] = unmatched_codes[:50]
        log_entry.metadata_json = md
        db.session.commit()

    logger.info(
        f"{tag} Cost import complete: "
        f"read={rows_read}, matched={rows_matched}, updated={rows_updated}, "
        f"skipped_blank_cost={rows_skipped_blank_cost}, skipped_no_code={rows_skipped_no_code}, "
        f"parse_errors={parse_errors}, unmatched={len(unmatched_codes)}"
    )
    return rows_updated


def disconnect_dropbox():
    cred = get_dropbox_credentials()
    if cred:
        cred.status = 'disconnected'
        cred.access_token = None
        cred.access_token_expires_at = None
        cred.last_error = None
        db.session.commit()
        logger.info("Dropbox disconnected")
    return True


def test_dropbox_connection():
    config = _get_config()
    if not config['file_path']:
        return {'success': False, 'error': 'DROPBOX_FILE_PATH not configured'}

    cred = get_dropbox_credentials()
    if not cred or cred.status != 'active':
        return {'success': False, 'error': 'Dropbox not connected'}

    try:
        metadata = get_dropbox_file_metadata(config['file_path'])
        return {
            'success': True,
            'file_name': metadata.get('name', ''),
            'file_size': metadata.get('size', 0),
            'modified': metadata.get('server_modified', ''),
            'revision': metadata.get('rev', ''),
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}
