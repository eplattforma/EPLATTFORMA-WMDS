"""Import suppliers from an XLSX file uploaded by an admin.

The expected sheet layout is the export from Powersoft365 with columns:

    Code | Detail Account No | Short Name | Name | Telephone | SMS | Balance

Only ``Code`` and ``Name`` are mandatory. Other columns may be empty.
"""
from __future__ import annotations

import logging
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

import openpyxl

from app import db
from models import Supplier

logger = logging.getLogger(__name__)


HEADER_ALIASES = {
    "code": "code",
    "supplier code": "code",
    "supplier_code": "code",
    "detail account no": "detail_account_no",
    "detail_account_no": "detail_account_no",
    "detail account": "detail_account_no",
    "short name": "short_name",
    "short_name": "short_name",
    "name": "name",
    "supplier name": "name",
    "telephone": "telephone",
    "phone": "telephone",
    "sms": "sms",
    "mobile": "sms",
    "balance": "balance",
}


def _norm(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _to_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return None


def parse_supplier_workbook(file_bytes: bytes) -> list[dict]:
    """Parse the uploaded workbook into a list of supplier dicts.

    Raises ``ValueError`` if the header row cannot be recognised.
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)

    try:
        header = next(rows)
    except StopIteration:
        raise ValueError("File is empty")

    col_map: dict[int, str] = {}
    for idx, cell in enumerate(header or []):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        target = HEADER_ALIASES.get(key)
        if target:
            col_map[idx] = target

    if "code" not in col_map.values() or "name" not in col_map.values():
        raise ValueError(
            "Could not find required columns 'Code' and 'Name' in the first row. "
            f"Found headers: {list(header) if header else []}"
        )

    parsed: list[dict] = []
    for row in rows:
        if row is None:
            continue
        record: dict[str, Any] = {}
        for idx, target in col_map.items():
            if idx >= len(row):
                continue
            value = row[idx]
            if target == "balance":
                record[target] = _to_decimal(value)
            else:
                record[target] = _norm(value)
        if not record.get("code") or not record.get("name"):
            continue
        parsed.append(record)
    return parsed


def upsert_suppliers(records: list[dict]) -> dict[str, int]:
    """Insert new suppliers and update existing ones (matched by code).

    Returns a summary of created/updated/unchanged counts.
    """
    created = updated = unchanged = 0
    existing_by_code = {s.code: s for s in Supplier.query.all()}

    # De-duplicate input by code (last occurrence wins) so a workbook with
    # repeated codes does not violate the unique constraint at commit time.
    deduped: dict[str, dict] = {}
    for r in records:
        deduped[r["code"]] = r

    for code, r in deduped.items():
        existing = existing_by_code.get(code)
        if existing is None:
            new_supplier = Supplier(
                code=code,
                detail_account_no=r.get("detail_account_no"),
                short_name=r.get("short_name"),
                name=r["name"],
                telephone=r.get("telephone"),
                sms=r.get("sms"),
                balance=r.get("balance"),
            )
            db.session.add(new_supplier)
            existing_by_code[code] = new_supplier
            created += 1
            continue

        changed = False
        for field in ("detail_account_no", "short_name", "name", "telephone", "sms", "balance"):
            new_val = r.get(field)
            if getattr(existing, field) != new_val:
                setattr(existing, field, new_val)
                changed = True
        if changed:
            updated += 1
        else:
            unchanged += 1

    db.session.commit()
    return {"created": created, "updated": updated, "unchanged": unchanged, "total": len(records)}
