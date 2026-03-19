"""
Admin Tools routes for database operations and maintenance
"""
import subprocess
import logging
from datetime import datetime, timezone
from flask import Blueprint, render_template, jsonify, request, Response, send_file, redirect, url_for
from flask_login import login_required, current_user
from flask import current_app as app
from app import db
import os
import openpyxl
from werkzeug.utils import secure_filename
from models import PostalCodeLookup

bp = Blueprint('admin_tools_custom', __name__, url_prefix='/admin/tools')
logger = logging.getLogger(__name__)

# Helper to check if user is admin
def is_admin():
    return current_user.is_authenticated and getattr(current_user, 'role', None) == 'admin'

@bp.route('/database-clone')
@login_required
def database_clone_page():
    """Show database clone tool page"""
    if not is_admin():
        return "Access denied", 403
    
    return render_template('admin_tools/database_clone.html')


@bp.route('/database-clone/execute', methods=['POST'])
@login_required
def execute_database_clone():
    """Execute database clone from production to development"""
    if not is_admin():
        return jsonify({"error": "Access denied"}), 403
    
    try:
        # Check environment variables
        database_url_prod = os.getenv('DATABASE_URL_PROD')
        database_url_dev = os.getenv('DATABASE_URL_DEV') or os.getenv('DATABASE_URL')
        
        if not database_url_prod:
            return jsonify({"error": "DATABASE_URL_PROD environment variable not set"}), 400
        
        if not database_url_dev:
            return jsonify({"error": "DATABASE_URL or DATABASE_URL_DEV not set"}), 400
        
        # Get confirmation
        if not request.get_json().get('confirmed'):
            return jsonify({"error": "Clone not confirmed"}), 400
        
        # Extract connection strings
        def parse_db_url(url):
            from urllib.parse import urlparse
            parsed = urlparse(url)
            return {
                'user': parsed.username,
                'password': parsed.password,
                'host': parsed.hostname,
                'port': parsed.port or 5432,
                'database': parsed.path.lstrip('/')
            }
        
        prod_config = parse_db_url(database_url_prod)
        dev_config = parse_db_url(database_url_dev)
        
        # Run pg_dump and psql
        dump_cmd = [
            'pg_dump',
            '-h', prod_config['host'],
            '-p', str(prod_config['port']),
            '-U', prod_config['user'],
            '-F', 'custom',
            prod_config['database']
        ]
        
        restore_cmd = [
            'pg_restore',
            '-h', dev_config['host'],
            '-p', str(dev_config['port']),
            '-U', dev_config['user'],
            '-d', dev_config['database'],
            '--no-owner',
            '--no-acl',
            '-j', '4'
        ]
        
        # Set password env vars
        env_vars = os.environ.copy()
        env_vars['PGPASSWORD'] = prod_config['password']
        
        logger.info("Starting database clone...")
        
        # Dump from prod
        dump_process = subprocess.Popen(dump_cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, env=env_vars)
        dump_output, dump_error = dump_process.communicate()
        
        if dump_process.returncode != 0:
            return jsonify({"error": f"Dump failed: {dump_error.decode()}"}), 500
        
        # Restore to dev
        env_vars['PGPASSWORD'] = dev_config['password']
        restore_process = subprocess.Popen(restore_cmd, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE, env=env_vars)
        restore_output, restore_error = restore_process.communicate(input=dump_output)
        
        if restore_process.returncode != 0:
            return jsonify({"error": f"Restore failed: {restore_error.decode()}"}), 500
        
        logger.info("Database clone completed successfully")
        return jsonify({"ok": True, "message": "Database cloned successfully"})
        
    except Exception as e:
        logger.error(f"Error cloning database: {e}")
        return jsonify({"error": str(e)}), 500



@bp.route('/postal-codes')
@login_required
def postal_codes_page():
    return redirect(url_for('admin_tools_custom.crm_classifications_settings'))


@bp.route('/postal-codes/import', methods=['POST'])
@login_required
def import_postal_codes():
    """Import postal codes from Excel file"""
    if not is_admin():
        return jsonify({"error": "Access denied"}), 403
    
    imported = 0
    duplicates = 0
    errors = []
    
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file provided"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({"error": "Only Excel files (.xlsx, .xls) are supported"}), 400
        
        # Load workbook
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        # Skip header row
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:  # Skip empty rows
                continue
            
            try:
                postcode = str(row[0]).strip()
                municipality = str(row[1]).strip() if row[1] else ""
                district = str(row[2]).strip() if row[2] else ""
                urban_rural = str(row[3]).strip() if row[3] else None
                
                # Try to add; if duplicate, mark it and continue
                lookup = PostalCodeLookup(
                    postcode=postcode,
                    municipality=municipality,
                    district=district,
                    urban_rural=urban_rural
                )
                db.session.add(lookup)
                imported += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
                # Clear pending adds for this row in case of error
                db.session.expunge(lookup) if 'lookup' in locals() else None
        
        # Attempt commit with proper error handling
        try:
            db.session.commit()
        except Exception as commit_err:
            db.session.rollback()
            # Check for duplicates and retry with skip-duplicates approach
            if "duplicate" in str(commit_err).lower() or "unique" in str(commit_err).lower():
                logger.info("Duplicate keys detected, retrying with skip-duplicates")
                db.session.expire_all()
                
                # Reload and check each postcode again
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                imported = 0
                duplicates = 0
                
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row[0]:
                        continue
                    
                    postcode = str(row[0]).strip()
                    # Check existence in DB before adding
                    if PostalCodeLookup.query.filter_by(postcode=postcode).first():
                        duplicates += 1
                        continue
                    
                    try:
                        municipality = str(row[1]).strip() if row[1] else ""
                        district = str(row[2]).strip() if row[2] else ""
                        urban_rural = str(row[3]).strip() if row[3] else None
                        
                        lookup = PostalCodeLookup(
                            postcode=postcode,
                            municipality=municipality,
                            district=district,
                            urban_rural=urban_rural
                        )
                        db.session.add(lookup)
                        imported += 1
                        
                        # Commit every 50 records to avoid lock timeouts
                        if imported % 50 == 0:
                            db.session.commit()
                    except Exception as e:
                        errors.append(f"Row {row_idx}: {str(e)}")
                
                db.session.commit()
            else:
                raise
        
        msg = f"Imported {imported} postal codes"
        if duplicates:
            msg += f", {duplicates} duplicates skipped"
        if errors:
            msg += f", {len(errors)} errors"
        
        return jsonify({
            "ok": True,
            "imported": imported,
            "duplicates": duplicates,
            "errors": errors[:10],  # Limit error messages
            "message": msg
        })
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error importing postal codes: {e}")
        return jsonify({"error": str(e)}), 500


@bp.route('/postal-codes/clear', methods=['POST'])
@login_required
def clear_postal_codes():
    """Clear all postal code data"""
    if not is_admin():
        return jsonify({"error": "Access denied"}), 403
    
    try:
        count = PostalCodeLookup.query.count()
        PostalCodeLookup.query.delete()
        db.session.commit()
        return jsonify({"ok": True, "cleared": count})
    except Exception as e:
        logger.error(f"Error clearing postal codes: {e}")
        return jsonify({"error": str(e)}), 500


@bp.route('/crm-classifications')
@login_required
def crm_classifications_settings():
    if not is_admin():
        return "Access denied", 403
    
    from models import Setting
    import json
    items = []
    allowed = Setting.get(db.session, "crm_customer_classifications", {})
    if isinstance(allowed, str):
        try:
            allowed = json.loads(allowed)
        except Exception:
            allowed = {}
    if isinstance(allowed, list):
        allowed = {name: "" for name in allowed}
    if not isinstance(allowed, dict):
        allowed = {}
    
    defaults = Setting.get(db.session, "crm_customer_classifications_defaults", [])
    if isinstance(defaults, str):
        try:
            defaults = json.loads(defaults)
        except Exception:
            defaults = []
    if not isinstance(defaults, list):
        defaults = []
    
    for name, meta in allowed.items():
        if isinstance(meta, dict):
            icon = meta.get("icon")
            include_in_review_ordering = meta.get("include_in_review_ordering", True)
        else:
            icon = meta or None
            include_in_review_ordering = True
        items.append({"name": name, "icon": icon, "is_default": name in defaults, "include_in_review_ordering": include_in_review_ordering})
    
    window_hours = Setting.get(db.session, "crm_order_window_hours", "48")
    if not isinstance(window_hours, str):
        window_hours = str(window_hours)
    
    anchor_time = Setting.get(db.session, "crm_delivery_anchor_time", "00:01")
    if not isinstance(anchor_time, str):
        anchor_time = str(anchor_time)
    
    close_hours = Setting.get(db.session, "crm_order_window_close_hours", "0")
    if not isinstance(close_hours, str):
        close_hours = str(close_hours)
    
    close_anchor_time = Setting.get(db.session, "crm_delivery_close_anchor_time", "00:01")
    if not isinstance(close_anchor_time, str):
        close_anchor_time = str(close_anchor_time)
    
    postal_code_count = PostalCodeLookup.query.count()
    return render_template('admin_tools/crm_classifications.html', items=items, window_hours=window_hours, anchor_time=anchor_time, close_hours=close_hours, close_anchor_time=close_anchor_time, postal_code_count=postal_code_count)


@bp.post('/crm-classifications/save')
@login_required
def save_crm_classifications():
    if not is_admin():
        return jsonify({"ok": False, "error": "Access denied"}), 403
    
    try:
        from models import Setting
        import json
        import re
        
        window_hours = (request.form.get('window_hours') or '48').strip()
        anchor_time = (request.form.get('anchor_time') or '00:01').strip()
        close_hours = (request.form.get('close_hours') or '0').strip()
        close_anchor_time = (request.form.get('close_anchor_time') or '00:01').strip()
        
        try:
            int(window_hours)
            int(close_hours)
        except ValueError:
            return jsonify({"ok": False, "error": "Window hours must be numbers"}), 400
        
        if not re.match(r'^([01][0-9]|2[0-3]):[0-5][0-9]$', anchor_time):
            return jsonify({"ok": False, "error": "Opening anchor time must be HH:MM format"}), 400
        
        if not re.match(r'^([01][0-9]|2[0-3]):[0-5][0-9]$', close_anchor_time):
            return jsonify({"ok": False, "error": "Closing anchor time must be HH:MM format"}), 400
        
        Setting.set(db.session, "crm_order_window_hours", window_hours)
        Setting.set(db.session, "crm_delivery_anchor_time", anchor_time)
        Setting.set(db.session, "crm_order_window_close_hours", close_hours)
        Setting.set(db.session, "crm_delivery_close_anchor_time", close_anchor_time)
        
        from routes_crm_dashboard import _normalize_classifications
        existing_raw = Setting.get(db.session, "crm_customer_classifications", "{}")
        existing = _normalize_classifications(existing_raw)
        
        data = request.form.get('data', '[]')
        items = json.loads(data) if data else []
        data_dict = {}
        defaults = []
        
        for item in items:
            name = (item.get('name') or '').strip()
            is_default = item.get('is_default', False)
            include_in_review_ordering = item.get('include_in_review_ordering', True)
            if not name:
                continue
            prev = existing.get(name, {"icon": None})
            data_dict[name] = {
                "icon": prev.get("icon"),
                "color": prev.get("color"),
                "sort_order": prev.get("sort_order"),
                "include_in_review_ordering": include_in_review_ordering,
            }
            if is_default:
                defaults.append(name)
        
        import base64, hashlib as _hashlib
        existing_b64 = Setting.get(db.session, 'crm_classification_images_b64', '{}')
        try:
            images_b64 = json.loads(existing_b64)
        except Exception:
            images_b64 = {}

        os.makedirs('static/crm-classification-images', exist_ok=True)
        from werkzeug.utils import secure_filename as _secure_filename
        for key in request.files:
            if key.startswith('file_'):
                idx = key.split('_', 1)[1]
                name_key = f'name_{idx}'
                cls_name = request.form.get(name_key, '').strip()
                if not cls_name or cls_name not in data_dict:
                    continue
                f = request.files[key]
                if f and f.filename:
                    ext = os.path.splitext(_secure_filename(f.filename))[1].lower()
                    if ext not in ('.jpg', '.jpeg', '.png', '.webp'):
                        continue
                    raw = f.read()
                    h = _hashlib.md5(raw).hexdigest()[:16]
                    safe_name = cls_name.replace(' ', '_')
                    filename = f"{safe_name}_{h}{ext}"
                    filepath = os.path.join('static/crm-classification-images', filename)
                    with open(filepath, 'wb') as fout:
                        fout.write(raw)
                    images_b64[filename] = base64.b64encode(raw).decode('ascii')
                    data_dict[cls_name]["icon"] = filename

        Setting.set(db.session, 'crm_classification_images_b64', json.dumps(images_b64))
        
        if not data_dict:
            return jsonify({"ok": False, "error": "At least one classification is required"}), 400
        
        Setting.set(db.session, "crm_customer_classifications", json.dumps(data_dict))
        Setting.set(db.session, "crm_customer_classifications_defaults", json.dumps(defaults))
        db.session.commit()
        
        return jsonify({"ok": True, "msg": f"Saved {len(data_dict)} classifications ({len(defaults)} as default) + ordering window settings"})
        
    except Exception as e:
        logger.error("Error saving classifications: %s", e)
        return jsonify({"ok": False, "error": str(e)}), 500


@bp.route('/crm-classifications/upload-icon', methods=['POST'])
@login_required
def upload_classification_icon():
    if not is_admin():
        return jsonify({"ok": False, "error": "Access denied"}), 403
    
    try:
        if 'file' not in request.files:
            return jsonify({"ok": False, "error": "No file"}), 400
        
        file = request.files['file']
        name = request.form.get('name', '').strip()
        
        if not file or not name:
            return jsonify({"ok": False, "error": "Missing file or name"}), 400
        
        if not file.filename.lower().endswith(('.jpg', '.jpeg', '.png', '.webp')):
            return jsonify({"ok": False, "error": "Only JPG/PNG/WebP files allowed"}), 400
        
        import base64, json, hashlib
        from werkzeug.utils import secure_filename as _secure_filename
        raw = file.read()
        h = hashlib.md5(raw).hexdigest()[:16]
        safe_name = name.replace(' ', '_')
        ext = os.path.splitext(_secure_filename(file.filename))[1].lower()
        if ext not in ('.jpg', '.jpeg', '.png', '.webp'):
            ext = '.png'
        filename = f"{safe_name}_{h}{ext}"
        b64 = base64.b64encode(raw).decode('ascii')

        from models import Setting
        existing_b64 = Setting.get(db.session, 'crm_classification_images_b64', '{}')
        try:
            images_dict = json.loads(existing_b64)
        except Exception:
            images_dict = {}
        images_dict[filename] = b64
        Setting.set(db.session, 'crm_classification_images_b64', json.dumps(images_dict))
        db.session.commit()

        os.makedirs('static/crm-classification-images', exist_ok=True)
        file.seek(0)
        filepath = os.path.join('static/crm-classification-images', filename)
        with open(filepath, 'wb') as fout:
            fout.write(raw)
        
        return jsonify({"ok": True, "filename": filename})
        
    except Exception as e:
        db.session.rollback()
        logger.error("Error uploading icon: %s", e)
        return jsonify({"ok": False, "error": str(e)}), 500


@bp.route('/crm-classification-image/<path:filename>')
@login_required
def serve_classification_image(filename):
    import json, base64, mimetypes, glob as globmod
    from io import BytesIO
    from werkzeug.utils import secure_filename as _secure_filename
    safe_name = _secure_filename(filename)
    if not safe_name or safe_name != filename:
        return Response('Invalid filename', status=400)

    img_dir = 'static/crm-classification-images'

    local_path = os.path.join(img_dir, safe_name)
    if os.path.exists(local_path):
        mime = mimetypes.guess_type(local_path)[0] or 'application/octet-stream'
        return send_file(local_path, mimetype=mime)

    from models import Setting
    raw_b64 = Setting.get(db.session, 'crm_classification_images_b64', '{}')
    try:
        images_dict = json.loads(raw_b64)
    except Exception:
        images_dict = {}
    b64_data = images_dict.get(filename)
    if b64_data:
        os.makedirs(img_dir, exist_ok=True)
        img_bytes = base64.b64decode(b64_data)
        with open(local_path, 'wb') as f:
            f.write(img_bytes)
        mime = mimetypes.guess_type(filename)[0] or 'application/octet-stream'
        return send_file(BytesIO(img_bytes), mimetype=mime)

    prefix = safe_name.rsplit('_', 1)[0] + '_' if '_' in safe_name else None
    if prefix:
        matches = globmod.glob(os.path.join(img_dir, prefix + '*'))
        if matches:
            fallback = matches[0]
            mime = mimetypes.guess_type(fallback)[0] or 'application/octet-stream'
            return send_file(fallback, mimetype=mime)
        for key in images_dict:
            if key.startswith(prefix):
                os.makedirs(img_dir, exist_ok=True)
                img_bytes = base64.b64decode(images_dict[key])
                fb_path = os.path.join(img_dir, key)
                with open(fb_path, 'wb') as f:
                    f.write(img_bytes)
                mime = mimetypes.guess_type(key)[0] or 'application/octet-stream'
                return send_file(BytesIO(img_bytes), mimetype=mime)

    return Response('Not found', status=404)


@bp.route('/crm-bulk-classify', methods=['GET', 'POST'])
@login_required
def crm_bulk_classify():
    """Bulk classify customers via CSV upload"""
    if not is_admin():
        return "Access denied", 403
    
    if request.method == 'POST':
        try:
            from models import CrmCustomerProfile
            import io
            import csv
            
            if 'file' not in request.files:
                return jsonify({"ok": False, "error": "No file uploaded"}), 400
            
            file = request.files['file']
            if not file or not file.filename.endswith('.csv'):
                return jsonify({"ok": False, "error": "CSV file required"}), 400
            
            stream = io.TextIOWrapper(file.stream, encoding='utf-8')
            reader = csv.DictReader(stream)
            
            updated = 0
            errors = []
            
            for row_idx, row in enumerate(reader, start=2):
                try:
                    code = (row.get('customer_code_365') or row.get('code') or '').strip()
                    classif = (row.get('classification') or '').strip()
                    
                    if not code:
                        continue
                    
                    prof = CrmCustomerProfile.query.get(code)
                    if not prof:
                        prof = CrmCustomerProfile(customer_code_365=code)
                        db.session.add(prof)
                    
                    prof.classification = classif or None
                    prof.updated_at = datetime.now(timezone.utc)
                    prof.updated_by = getattr(current_user, "username", None)
                    updated += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")
            
            db.session.commit()
            msg = f"Classified {updated} customers"
            if errors:
                msg += f" ({len(errors)} errors)"
            return jsonify({"ok": True, "updated": updated, "errors": errors, "msg": msg})
        except Exception as e:
            logger.error("Error bulk classifying: %s", e)
            return jsonify({"ok": False, "error": str(e)}), 500

    return render_template('admin_tools/crm_bulk_classify.html')




@bp.route('/magento-last-login-data')
@login_required
def magento_last_login_data_page():
    if not is_admin():
        return "Forbidden", 403

    from models import MagentoCustomerLastLoginCurrent
    page = request.args.get('page', 1, type=int)
    per_page = 50
    q = request.args.get('q', '').strip()

    query = MagentoCustomerLastLoginCurrent.query
    if q:
        like = f"%{q}%"
        query = query.filter(
            db.or_(
                MagentoCustomerLastLoginCurrent.customer_code_365.ilike(like),
                MagentoCustomerLastLoginCurrent.email.ilike(like),
                MagentoCustomerLastLoginCurrent.first_name.ilike(like),
                MagentoCustomerLastLoginCurrent.last_name.ilike(like),
            )
        )

    total = query.count()
    rows = (query
            .order_by(MagentoCustomerLastLoginCurrent.last_login_at.desc().nullslast())
            .offset((page - 1) * per_page)
            .limit(per_page)
            .all())
    total_pages = max(1, (total + per_page - 1) // per_page)

    return render_template('admin_tools/magento_last_login_data.html',
                           rows=rows, page=page, total_pages=total_pages,
                           total=total, q=q)


@bp.route('/magento-last-login-sample')
@login_required
def magento_last_login_sample():
    if not is_admin():
        return jsonify({"ok": False, "error": "Forbidden"}), 403

    from models import MagentoCustomerLastLoginCurrent
    rows = (MagentoCustomerLastLoginCurrent.query
            .order_by(MagentoCustomerLastLoginCurrent.last_login_at.desc().nullslast())
            .limit(10).all())

    return jsonify([{
        "customer_code_365": r.customer_code_365,
        "magento_customer_id": r.magento_customer_id,
        "last_login_at_utc": r.last_login_at.isoformat() if r.last_login_at else None,
        "email": r.email,
        "source": r.source_filename,
    } for r in rows])
